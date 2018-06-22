# -*- coding: utf-8 -*-
"""This is the main file for converting excel to csv files, processing those
csv files, and merging them for upload to calfreshdb

Output:
    writes files to the <outpath> defined after the import statements

"""

from csv import writer
from datetime import datetime
from os import walk, remove, makedirs
from os.path import join, exists
from shutil import move
import ConfigParser
import logging.config

from openpyxl import load_workbook
from xlrd import open_workbook
import pandas as pd

from file_factory import initialize

config = ConfigParser.RawConfigParser()
config.read('/etc/calfresh/calfresh.conf')

logging.config.fileConfig(config.get('filepaths', 'config'))
logger = logging.getLogger('worker')

INPATH = config.get('filepaths', 'data')
now = datetime.now()
OUTPATH = '/etc/calfresh/{}_{}_{}'.format(now.month, now.day, now.year)


class Worker(object):
    def __init__(self, table):
        """The worker performs the data cleaning and standardization
        Args:
            table (str): the table type the data to process belongs to

        Returns:
            table (str): the table, so the data loader knows what to load

        """
        self.table = table
        if not exists(OUTPATH):
            makedirs(OUTPATH)

    def work(self):
        """Do the needful: convert the files, run the factories, merge the output"""
        # convert excel files to csv
        self.excelToCSV()
        paths = self.getCSVInput()
        self.removeJunkFiles(paths)

        # run the factories for processing
        paths = self.getCSVInput()
        self.runFactories(paths)

        # merge the files
        paths = self.getCSVOutput()
        if self.table == 'tbl_data_dashboard':
            self.redistributeDataDashboardFiles(paths)
        else:
            self.mergeForUploading(paths)

        return OUTPATH

    def getCSVInput(self):
        """Search directories for unprocessed csv files

        Returns:
            paths (list of dicts): contains the dict objects representing
            the path, source, and filename of each csv file to be processed

        """
        paths = []
        for root, dirs, files in walk(INPATH):
            for name in files:
                if name.endswith('.csv') and root.split('/')[5] == 'csv_in':
                    if root.split('/')[4] == self.table:
                        paths.append({
                                     'path': join(root, name),
                                     'source': root.split('/')[4],
                                     'filename': name
                                     })
        return paths

    def getCSVOutput(self):
        """Search directories for processed csv files

        Returns:
            paths (list of dicts): contains the dict objects representing
            the path, source, and filename of each csv file to be merged

        """
        paths = []
        for root, dirs, files in walk(INPATH):
            for name in files:
                if name.endswith('.csv') and root.split('/')[5] == 'csv_out':
                    if root.split('/')[4] == self.table:
                        paths.append({
                                     'path': join(root, name),
                                     'source': root.split('/')[4],
                                     'filename': name
                                     })
        return paths

    def getExcelFiles(self):
        """Search directories for excel files

        Args:
            inpath (str): the path to recursively search for excel files

        Returns:
            paths (list of dicts): contains the dict objects representing
            the path, source, and filename of each excel file to be converted

        """
        paths = []
        for root, dirs, files in walk(INPATH):
            for name in files:
                if '.xls' in name:
                    if root.split('/')[4] == self.table:
                        paths.append({
                                     'path': join(root, name),
                                     'source': root.split('/')[4],
                                     'filename': name
                                     })
        return paths

    def convertExcelFile(self, item):
        """Convert excel files ending in .xls and .xlsx

        Args:
            item (dict): dict with keynames path, source, and filename

        Output:
            writes each tab of the file out to csv

        """
        workbook = open_workbook(item['path'])

        filename = self.stripFilename(item['filename'])

        worksheets = workbook.sheet_names()

        for name in worksheets:
            sheet = workbook.sheet_by_name(name)
            with open(
                join(
                    INPATH,
                    item['source'],
                    'csv_in',
                    filename + '-' + name + '.csv',
                ),
                'wb'
            ) as handle:
                author = writer(handle)
                for row in xrange(sheet.nrows):
                    author.writerow([
                        unicode(value).encode('utf-8') for value in sheet.row_values(row)
                    ])

    def stripFilename(self, filename):
        """Removes the suffix from excel files

        Args:
            filename (str): filename ending in .xls or .xlsx

        Returns:
            substring of filename ending before '.xls'

        """
        return filename.split('.xls')[0]

    def convertNewExcelFile(self, item):
        """Convert excel files ending in .xlsx, representing excel versions >2010

        Args:
            item (dict): dict with keys path, source, and filename of .xls excel files

        Output:
            writes each tab of the file out to csv

        """
        workbook = load_workbook(item['path'])

        filename = self.stripFilename(item['filename'])

        worksheets = workbook.sheetnames

        for name in worksheets:
            sheet = workbook[name]
            with open(
                    join(
                        INPATH,
                        item['source'],
                        'csv_in',
                        filename + '-' + name + '.csv'
                    ),
                    'wb'
            ) as handle:
                author = writer(handle)
                author.writerows([
                    unicode(value).encode('utf-8') for value in sheet.iter_rows()
                ])

    def excelToCSV(self):
        """Convert all excel files to csvs in the source directories"""
        paths = self.getExcelFiles()
        for item in paths:
            if item['source'] != self.table:
                continue

            logger.info('converting: %s', item['filename'])
            self.convertExcelFile(item)

    def redistributeDataDashboardFiles(self, paths):
        for path in paths:
            if path['filename'] == 'CFDashboard-Annual.csv':
                move(
                    path['path'],
                    join(OUTPATH, path['filename'])
                )
            elif path['filename'] == 'CFDashboard-Quarterly.csv':
                move(
                    path['path'],
                    join(OUTPATH, path['filename'])
                )
            elif path['filename'] == 'CFDashboard-Every_Mth.csv':
                move(
                    path['path'],
                    join(OUTPATH, path['filename'])
                )
            elif path['filename'] == 'CFDashboard-Every_3_Mth.csv':
                move(
                    path['path'],
                    join(OUTPATH, path['filename'])
                )
            elif path['filename'] == 'CFDashboard-PRI_Raw.csv':
                move(
                    path['path'],
                    join(OUTPATH, path['filename'])
                )

    def removeJunkFiles(self, paths):
        """Remove files that don't contain relevant data

        Args:
            paths (list of str): the file paths to search for junk files

        """
        for path in paths:
            if path['source'] == 'tbl_cf15':
                continue

            elif 'DataDictionary' in path['filename']:
                remove(path['path'])

            elif path['source'] == 'tbl_cf296':
                if 'Statewide' in path['filename']:
                    remove(path['path'])
                elif 'Release Summary' in path['filename']:
                    remove(path['path'])
                elif 'Report View' in path['filename']:
                    remove(path['path'])

            elif path['source'] == 'tbl_churn_data':
                continue

            elif path['source'] == 'tbl_data_dashboard':
                if 'Trend' in path['filename']:
                    remove(path['path'])
                elif 'Updates' in path['filename']:
                    remove(path['path'])
                elif 'PRI_eval' in path['filename']:
                    remove(path['path'])
                elif 'Pivot' in path['filename']:
                    remove(path['path'])
                elif 'Main' in path['filename']:
                    remove(path['path'])
                elif 'Geomap' in path['filename']:
                    remove(path['path'])
                elif 'Dual_Part' in path['filename']:
                    remove(path['path'])
                elif 'Tiered' in path['filename']:
                    remove(path['path'])
                elif '_US' in path['filename']:
                    remove(path['path'])

            elif path['source'] == 'tbl_dfa256':
                if 'Statewide' in path['filename']:
                    remove(path['path'])
                elif 'Release Summary' in path['filename']:
                    remove(path['path'])
                elif 'Report View' in path['filename']:
                    remove(path['path'])

            elif path['source'] == 'tbl_dfa296':
                if 'Statewide' in path['filename']:
                    remove(path['path'])
                elif 'Release Summary' in path['filename']:
                    remove(path['path'])

            elif path['source'] == 'tbl_dfa296x':
                if 'Statewide' in path['filename']:
                    remove(path['path'])
                elif 'Release Summary' in path['filename']:
                    remove(path['path'])

            elif path['source'] == 'tbl_dfa358f':
                if 'Statewide' in path['filename']:
                    remove(path['path'])
                elif 'Release Summary' in path['filename']:
                    remove(path['path'])

            elif path['source'] == 'tbl_dfa358s':
                if 'Statewide' in path['filename']:
                    remove(path['path'])
                elif 'Release Summary' in path['filename']:
                    remove(path['path'])

            elif path['source'] == 'tbl_stat47':
                if 'Statewide' in path['filename']:
                    remove(path['path'])
                elif 'Release Summary' in path['filename']:
                    remove(path['path'])

            elif path['source'] == 'tbl_stat48':
                if '0.csv' in path['filename']:
                    remove(path['path'])

    def runFactories(self, paths):
        """Process all the csv files in the directories specified

        Args:
            paths (list of str): all the file paths to process in the factories

        """
        for item in paths:
            if item['source'] != self.table:
                continue

            logger.info('Processing file: %s', item['filename'])
            factory = initialize(item)
            if item['filename'] in [
                'CFDashboard-Annual.csv',
                'CFDashboard-Quarterly.csv',
                'CFDashboard-Every_Mth.csv',
                'CFDashboard-Every_3_Mth.csv',
                'CFDashboard-PRI_Raw.csv',
            ]:
                if item['filename'] == 'CFDashboard-Annual.csv':
                    factory.df.to_csv(
                        join(
                            INPATH,
                            item['source'],
                            'csv_out',
                            'tbl_data_dashboard_annual.csv',
                        ), index=False)
                elif item['filename'] == 'CFDashboard-Quarterly.csv':
                    factory.df.to_csv(
                        join(
                            INPATH,
                            item['source'],
                            'csv_out',
                            'tbl_data_dashboard_quarterly.csv',
                        ), index=False)
                elif item['filename'] == 'CFDashboard-Every_Mth.csv':
                    factory.df.to_csv(
                        join(
                            INPATH,
                            item['source'],
                            'csv_out',
                            'tbl_data_dashboard_monthly.csv',
                        ), index=False)
                elif item['filename'] == 'CFDashboard-Every_3_Mth.csv':
                    factory.df.to_csv(
                        join(
                            INPATH,
                            item['source'],
                            'csv_out',
                            'tbl_data_dashboard_3mth.csv',
                        ),
                        index=False
                    )
                elif item['filename'] == 'CFDashboard-PRI_Raw.csv':
                    factory.df.to_csv(
                        join(
                            INPATH,
                            item['source'],
                            'csv_out',
                            'tbl_data_dashboard_pri_raw.csv',
                        ),
                        index=False
                    )
            else:
                factory.df.to_csv(
                    join(
                        INPATH,
                        item['source'],
                        'csv_out',
                        item['filename'],
                    ),
                    index=False
                )

    def mergeForUploading(self, paths):
        """Merge all the csv files in the directories specified for uploading to the database

        Args:
            paths (list of dicts): each dict has a path, source, and filename
            to be used to read in csv files and then merge them according
            to their source

        Output:
            writes merged csv files out to the outpath

        """
        sibling = paths[0]['source']
        old = pd.read_csv(paths[0]['path'])

        for item in paths[1:]:

            new = pd.read_csv(item['path'])

            if item['source'] == sibling:
                old = pd.merge(old, new, how='outer')
            else:
                old.to_csv(join(OUTPATH, sibling + '.csv'), index=False)
                old = new
                sibling = item['source']

        old.to_csv(join(OUTPATH, sibling + '.csv'), index=False)
        logger.info('Merged files for %s', sibling)

        if self.table in ['tbl_dfa358f', 'tbl_dfa358s']:
            self.combine358FandS()

    def combine358FandS(self):
        """Combines the tbl_dfa358f and tbl_dfa358s files for uploading

        Args:
            outpath (str): path to the directory of the tbl_dfa358f and tbl_dfa358s files

        Output:
            csv file named tbl_dfa358tot in the same outpath directory

        Raises:
            ValueError: If the combined file is empty, something went wrong

        """
        df1 = pd.read_csv(join(OUTPATH, 'tbl_dfa358f.csv'))
        df2 = pd.read_csv(join(OUTPATH, 'tbl_dfa358s.csv'))

        df3 = df1.append(df2, ignore_index=True)
        df3 = df3.groupby(
            [df3.county, df3.fulldate, df3.year, df3.quarter, df3.month]
        ).sum()

        if df3.empty:
            raise ValueError

        df3.to_csv(join(OUTPATH, 'tbl_dfa358tot.csv'))
