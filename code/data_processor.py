"""data_processor is the main file for converting excel to csv files, processing those csv files, and merging them for upload to calfreshdb

Args:
    <none> : if nothing is specified, run the entire program on all directories

    <directory> : if no flags are specified, run the entire program on this directory of files, where directory is one of the following: ['tbl_cf15','tbl_cf296','tbl_churn_data','tbl_data_dashboard','tbl_dfa256','tbl_dfa296','tbl_dfa296x','tbl_dfa358f','tbl_dfa358s','tbl_stat47','tbl_stat48']


    the flags below run on all directories if no directories are specified, otherwise they run only on those passed in. multiple flags must be passed in order to run multiple processor functions, otherwise only the process specified by the flag will be run

    -etoc : convert excel files to csv

    -r : run the processor

    -m : merge csv files for uploading to database

    --help : print this docstring and end the program

Output:
    writes files to the <outpath> defined after the import statements

TODOs:
    Finish and clean up documentation 
    Build tests
    Refactor as necessary to write better tests
"""


from os import walk
from os.path import join, split
from xlrd import open_workbook
import sys
from csv import writer
from shutil import move

sys.path.append('/Users/Home/anaconda2/lib/python2.7/site-packages/')
from openpyxl import load_workbook
import pandas as pd

from constants import constants
from file_factory import initialize
     

INPATH = '/Users/Home/Google_Drive/Work/CalFresh/Data/Revised'
OUTPATH = '/Users/Home/Google_Drive/Work/CalFresh/Data/8_1_17_update'
DIRECTORIES = ['tbl_cf15','tbl_cf296','tbl_churn_data','tbl_data_dashboard','tbl_dfa256','tbl_dfa296','tbl_dfa296x','tbl_dfa358f','tbl_dfa358s','tbl_dfa358tot','tbl_stat47','tbl_stat48']
FLAGS = ['-etoc','-r','-m']


def getCSVInput():
    """search directories for unprocessed csv files

    Args:
        inpath (str): the path to recursively search for csv files

    Returns:
        paths (list of dicts): contains the dict objects representing the path, source, and filename of each csv file to be processed
    """

    paths = []
    for root, dirs, files in walk(INPATH):
        for name in files:
            if name.endswith('.csv') and root.split('/')[9] == 'csv_in':
                if root.split('/')[8] in DIRECTORIES:
                    paths.append({
                                 'path': join(root, name),
                                 'source': root.split('/')[8],
                                 'filename': name
                                 })
    return paths


def getCSVOutput():
    """search directories for processed csv files

    Args:
        inpath (str): the path to recursively search for csv files

    Returns:
        paths (list of dicts): contains the dict objects representing the path, source, and filename of each csv file to be merged
    """

    paths = []
    for root, dirs, files in walk(INPATH):
        for name in files:
            if name.endswith('.csv') and root.split('/')[9] == 'csv_out':
                if root.split('/')[8] in DIRECTORIES:
                    paths.append({
                                 'path': join(root, name),
                                 'source': root.split('/')[8],
                                 'filename': name
                                 })
    return paths


def getExcelFiles():
    """search directories for excel files

    Args:
        inpath (str): the path to recursively search for excel files

    Returns:
        paths (list of dicts): contains the dict objects representing the path, source, and filename of each excel file to be converted
    """
    paths = []
    for root, dirs, files in walk(INPATH):
        for name in files:
            if '.xls' in name:
                if root.split('/')[8] in DIRECTORIES:
                    paths.append({
                                 'path': join(root, name),
                                 'source': root.split('/')[8],
                                 'filename': name
                                 })
    return paths


def convertExcelFile(item):
    """convert excel files ending in .xls and .xlsx

    Args:
        item (dict): dict with keynames path, source, and filename

    Output:
        writes each tab of the file out to csv
    """

    workbook = open_workbook(item['path'])

    filename = stripFilename(item['filename'])

    worksheets = workbook.sheet_names()

    for name in worksheets:
        sheet = workbook.sheet_by_name(name)
        with open(join(INPATH, item['source'], 'csv_in', filename+'-'+name+'.csv'), 'wb') as handle:
            author = writer(handle)
            for row in xrange(sheet.nrows):
                author.writerow([unicode(value).encode('utf-8') for value in sheet.row_values(row)])


def stripFilename(filename):
    """removes the suffix from excel files

    Args:
        filename (str): filename ending in .xls or .xlsx

    Returns:
        substring of filename ending before '.xls'
    """
    return filename.split('.xls')[0]


def convertNewExcelFile(item):
    """convert excel files ending in .xlsx, representing excel versions >2010

    Args:
        item (dict): dict with keys path, source, and filename of excel files ending in .xls

    Output:
        writes each tab of the file out to csv
    """

    workbook = load_workbook(item['path'])

    filename = stripFilename(item['filename'])

    worksheets = workbook.sheetnames

    for name in worksheets:
        sheet = workbook[name]
        with open(join(INPATH, item['source'], 'csv_in', filename+'-'+name+'.csv'), 'wb') as handle:
            author = writer(handle)
            author.writerows([unicode(value).encode('utf-8') for value in sheet.iter_rows()])


def excelToCSV():
    """convert all excel files to csvs in the source directories"""

    paths = getExcelFiles()
    for item in paths:
        if item['source'] not in DIRECTORIES: 
            continue

        print 'converting:', item['filename']
        convertExcelFile(item)



def removeJunkFiles(paths):
    """remove files known to not to contain data"""

    junkfolder = '/Users/Home/Google_Drive/Work/CalFresh/Data/Junk'
    for path in paths:
        if path['source'] == 'tbl_cf15':
            continue


        elif path['source'] == 'tbl_cf296':
            if 'Statewide' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Release Summary' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))


        elif path['source'] == 'tbl_churn_data':
            continue


        elif path['source'] == 'tbl_data_dashboard':
            if 'Trend' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Updates' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Quarterly' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'PRI' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Pivot' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Main' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Every' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Annual' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif '0.csv' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif '2.csv' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif '3.csv' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))


        elif path['source'] == 'tbl_dfa256':
            if 'Statewide' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Release Summary' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))


        elif path['source'] == 'tbl_dfa296':
            if 'Statewide' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Release Summary' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))


        elif path['source'] == 'tbl_dfa296x':
            if 'Statewide' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Release Summary' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            

        elif path['source'] == 'tbl_dfa358f':
            if 'Statewide' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Release Summary' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            

        elif path['source'] == 'tbl_dfa358s':
            if 'Statewide' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Release Summary' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
        

        elif path['source'] == 'tbl_stat47':
            if 'Statewide' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            elif 'Release Summary' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))
            

        elif path['source'] == 'tbl_stat48':
            if '0.csv' in path['filename']:
                move(path['path'], join(junkfolder, path['filename']))


def runProcessor(paths):
    """process all the csv files in the directories specified"""

    for item in paths:
        if item['source'] not in DIRECTORIES: 
            continue

        print 'input:', item['filename']
        factory = initialize(item)
        factory.df.to_csv(join(INPATH, item['source'], 'csv_out', item['filename']), index=False)


def mergeForUploading(paths):
    """merge all the csv files in the directories specified for uploading to the database

    Args:
        paths (list of dicts): each dict has a path, source, and filename to be used to read in csv files and then merge them according to their source

    Output:
        writes merged csv files out to the outpath
    """

    sibling = paths[0]['source']
    old = pd.read_csv(paths[0]['path'])

    for item in paths[1:]: 

        new = pd.read_csv(item['path'])

        if item['source'] == sibling:
            old = pd.merge(old, new, how='outer')
        else:
            old.to_csv(join(OUTPATH, sibling + '.csv'), index=False)
            print 'output:', sibling
            old = new
            sibling = item['source']

    old.to_csv(join(OUTPATH, sibling + '.csv'), index=False)
    print 'output:', sibling

    for tbl in DIRECTORIES:
        if tbl in ['tbl_dfa358f','tbl_dfa358s']:
            combine358FandS()
            print 'output: tbl_dfa358tot'


def combine358FandS():
    """combines the tbl_dfa358f and tbl_dfa358s files for uploading to the database

    Args:
        outpath (str): path to the directory of the tbl_dfa358f and tbl_dfa358s files

    Output:
        csv file named tbl_dfa358tot in the same outpath directory

    Raises:
        ValueError: If the combined file is empty, something went wrong
    """
    
    df1 = pd.read_csv(join(OUTPATH, 'tbl_dfa358f.csv'))
    df2 = pd.read_csv(join(OUTPATH, 'tbl_dfa358s.csv'))

    df3 = df1.append(df2, ignore_index=True)
    df3 = df3.groupby([df3.county, df3.fulldate, df3.year, df3.quarter, df3.month]).sum()

    if df3.empty: raise ValueError

    df3.to_csv(join(OUTPATH,'tbl_dfa358tot.csv'))




if __name__ == '__main__':

    """parse system arguments"""
    specifiedDirs = []
    specifiedFlags = []
    for arg in sys.argv[1:]:
        if arg == '--help':
            print __doc__
            DIRECTORIES = []
            specifiedDirs = []
            FLAGS = []
            specifiedFlags = []
            break

        # if any directories are specified to be processed, only run those
        if arg in DIRECTORIES:
            specifiedDirs.append(arg)

        # if any flags are specified, only run their associated processes
        elif arg in FLAGS:
            specifiedFlags.append(arg)

        # raise an error on all other argvs supplied, except the filename, which gets skipped
        else:
            print '\nUnknown arg passed to program:', arg, '\n'
            print 'Directories allowed:', DIRECTORIES, '\n'
            print 'Flags allowed:', FLAGS, '\n'
            raise ValueError

    if specifiedDirs:
        DIRECTORIES = specifiedDirs

    if specifiedFlags:
        FLAGS = specifiedFlags
    
    if '-etoc' in FLAGS:
        excelToCSV()
        paths = getCSVInput()
        removeJunkFiles(paths)

    if '-r' in FLAGS:
        paths = getCSVInput()
        runProcessor(paths)

    if '-m' in FLAGS:
        paths = getCSVOutput()
        mergeForUploading(paths)


























